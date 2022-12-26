from openpyxl import load_workbook
import openpyxl
from datetime import datetime
import pyrebase
import time
import smtplib
import ssl

#declare the firebase configration

firebaseConfig = {
  'apiKey': "AIzaSyDAyEfudZKQ7zZrKcHt4JNh3zRvcJTGGhM",
  'authDomain': "yu-attendance.firebaseapp.com",
  'projectId':"yu-attendance",
  'storageBucket': "yu-attendance.appspot.com",
  'messagingSenderId':"1053945043332",
  'appId': "1:1053945043332:web:6a3fb660823e6e424225d1",
  'measurementId': "G-XJFT9HJ8TW",
'databaseURL' : ""
}

# initialize the firebase storage (connect)

firebase = pyrebase.initialize_app(firebaseConfig)
storage=firebase.storage()


starttime=time.time()



class attendance :

    #initial the excel sheet name
    def __init__(self,sectionName,sectionId,sectionNumer):
        self.sectionName=sectionName
        self.sectionId=sectionId
        self.sectionNumber=sectionNumer



    #this function Calculates how many absences on student , the argument is student id
    def check_stdabsence(self,std_id):
        workBook=load_workbook(filename=self.sectionName+self.sectionId+self.sectionNumber+".xlsx")
        sheet=workBook.active
        abscen_counter=0
        dates=""


        #for loop to find the student id in excel sheet
        for i in range(1,  sheet.max_row+1):

            # statement to check if (i) value match the student or not
            if sheet.cell(row=i, column=1).value == std_id:

                #if founded create inner loop to retrive data from the column
                for j in range(4, sheet.max_column+1):
                    founded_target = sheet.cell(row=i, column=j).value

                    #statement to check if the cell is not empty add to the string the cell value and increamnt the counter by one
                    if founded_target != None:
                        abscen_counter+=1
                        dates+=str(founded_target)+"  -  "

        return (dates +"\ntotal:" + str(abscen_counter))


    #this function to mark the student as absent in excel sheet
    def mark_attend(self, std_id):
        workBook = load_workbook(filename=self.sectionName+self.sectionId+self.sectionNumber+".xlsx")
        sheet = workBook.active

        # for loop to find the student id in excel sheet
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == std_id:

                column_target = sheet[i]    # storing the list of the cells value

                # (<Cell 'Sheet1'.A6>, <Cell 'Sheet1'.B6>)
                numberstd=sheet.cell(row=i,column=1).value
                email=sheet.cell(row=i,column=3).value
                #self.send(numberstd, email)

                if  column_target[-1].value != None:
                    #print("col  if ", column_target)
                    cell_target = str( column_target[-1])   #<Cell 'Sheet1'.B6>
                    cell_target = cell_target.split(".")    #["<Cell 'Sheet1","B6>"]
                    cell_target = cell_target[1]   #"B6>"
                    cell_target = cell_target[:len(cell_target) - 1]  #"B6"

                    next_cell = chr(ord(cell_target[0]) + 1)  #C
                    empty_cell = next_cell + cell_target[1:]   # "c" +"6" =c6
                    now = datetime.now()
                    dtString = now.strftime('%d/%m/%y')
                    sheet[empty_cell] = dtString



                    workBook.save(filename=str(self.sectionName) + str(self.sectionId) + str(self.sectionNumber)+".xlsx")
                    path_on_cloud="sheets/"+str(self.sectionName) + str(self.sectionId) + str(self.sectionNumber)+".xlsx"
                    storage.child(path_on_cloud).put(str(self.sectionName) + str(self.sectionId) + str(self.sectionNumber)+".xlsx")


                    break

                else:
                    #print("col  else ", column_target)
                    for i in column_target:

                        if i.value == None:
                           cell_target = (str(i))
                           cell_target = cell_target.split(".")
                           cell_target = cell_target[1]
                           empty_cell = cell_target[:len(cell_target) - 1]
                           now = datetime.now()
                           dtString = now.strftime('%d/%m/%y')
                           sheet[empty_cell] = dtString

                           workBook.save(filename=str(self.sectionName) + str(self.sectionId) + str(self.sectionNumber) + ".xlsx")
                           path_on_cloud = "sheets/" + str(self.sectionName) + str(self.sectionId) + str(self.sectionNumber) + ".xlsx"
                           storage.child(path_on_cloud).put(str(self.sectionName) + str(self.sectionId) + str(self.sectionNumber) + ".xlsx")
                           break



    def send(self, std_num, std_email):
        port = 465  # For SSL
        password = "oqbekykovygbmkgp"

        # returns a new context with secure default settings.
        context = ssl.create_default_context()
        sender_email = "yu.attend@gmail.com"

        section_name =str(self.sectionName) + str(self.sectionId)
        absent_number = str(self.check_stdabsence(std_num))
        absent_number = absent_number.split("total:")
        absent_number = str(absent_number[1])
        absent_number =(int(absent_number) + 1)

        if absent_number >6:
            status="total  number of absences :" +str(absent_number) +" status: deprived from course due to absence"

        else:
            status="total  number of absences :" +str(absent_number) + " status: in progress" +"\n"+"remaining :"+str(7-absent_number)


        message = """
        Subject: You Have been recorded 


        You have been recorded as absent from the course:  """ +str(section_name)+" \n"+status

        with smtplib.SMTP_SSL("smtp.gmail.com", port, context=context) as server:

            #TODO: login and send email

            server.login(sender_email, password)
            server.sendmail(sender_email, std_email, message)



